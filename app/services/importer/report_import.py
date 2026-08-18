from __future__ import annotations

import csv
import math
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import Product, Shop, Sku, SkuDailyMetric
from app.services.importer.report_reader import detect_columns


@dataclass(slots=True)
class ImportSummary:
    rows_total: int = 0
    rows_imported: int = 0
    rows_skipped: int = 0
    products_created: int = 0
    skus_created: int = 0
    metrics_created: int = 0
    metrics_updated: int = 0
    min_date: date | None = None
    max_date: date | None = None

    def as_dict(self) -> dict[str, object]:
        return {
            "rows_total": self.rows_total,
            "rows_imported": self.rows_imported,
            "rows_skipped": self.rows_skipped,
            "products_created": self.products_created,
            "skus_created": self.skus_created,
            "metrics_created": self.metrics_created,
            "metrics_updated": self.metrics_updated,
            "date_from": self.min_date.isoformat() if self.min_date else None,
            "date_to": self.max_date.isoformat() if self.max_date else None,
        }


def _iter_rows(path: Path) -> tuple[list[str], Iterable[dict[str, Any]]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        handle = path.open("r", encoding="utf-8-sig", newline="")
        reader = csv.DictReader(handle)
        headers = [str(item or "").strip() for item in (reader.fieldnames or [])]

        def rows() -> Iterable[dict[str, Any]]:
            try:
                yield from reader
            finally:
                handle.close()

        return headers, rows()

    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(iterator, ())]

        def rows() -> Iterable[dict[str, Any]]:
            try:
                for values in iterator:
                    yield {header: value for header, value in zip(headers, values)}
            finally:
                workbook.close()

        return headers, rows()

    raise ValueError("仅支持 CSV、XLSX、XLSM 报表")


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return Decimal(int(value))
    text = str(value).strip().replace(",", "").replace("￥", "").replace("¥", "")
    if not text or text in {"-", "--", "N/A", "n/a"}:
        return None
    if text.endswith("%"):
        text = text[:-1]
        try:
            return Decimal(text) / Decimal(100)
        except InvalidOperation:
            return None
    try:
        result = Decimal(text)
    except InvalidOperation:
        return None
    if not result.is_finite():
        return None
    return result


def _integer(value: Any) -> int | None:
    number = _decimal(value)
    if number is None:
        return None
    try:
        result = int(number)
    except (ValueError, OverflowError):
        return None
    return result


def _date(value: Any) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if math.isfinite(float(value)) and 20_000 <= float(value) <= 80_000:
            return date.fromordinal(date(1899, 12, 30).toordinal() + int(value))
    text = str(value).strip()
    for fmt in (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y%m%d",
    ):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _lookup(row: dict[str, Any], mapping: dict[str, str], field: str) -> Any:
    column = mapping.get(field)
    return row.get(column) if column else None


def _get_or_create_product(
    session: Session,
    *,
    shop_id: int,
    goods_id: str,
    summary: ImportSummary,
) -> Product:
    product = session.scalar(
        select(Product).where(Product.shop_id == shop_id, Product.platform_goods_id == goods_id)
    )
    if product is not None:
        return product
    product = Product(
        shop_id=shop_id,
        platform_goods_id=goods_id,
        title=f"报表商品 {goods_id}",
        status="report_import",
    )
    session.add(product)
    session.flush()
    summary.products_created += 1
    return product


def _get_or_create_sku(
    session: Session,
    *,
    product: Product,
    platform_sku_id: str,
    summary: ImportSummary,
) -> Sku:
    sku = session.scalar(
        select(Sku).where(
            Sku.product_id == product.id,
            Sku.platform_sku_id == platform_sku_id,
        )
    )
    if sku is not None:
        return sku
    sku = Sku(
        product_id=product.id,
        platform_sku_id=platform_sku_id,
        sku_name=f"SKU {platform_sku_id}",
        status="report_import",
    )
    session.add(sku)
    session.flush()
    summary.skus_created += 1
    return sku


METRIC_FIELDS = {
    "impression": _integer,
    "visitors": _integer,
    "clicks": _integer,
    "order_count": _integer,
    "sales_qty": _integer,
    "gmv": _decimal,
    "refund_count": _integer,
    "refund_amount": _decimal,
    "ad_cost": _decimal,
    "ad_clicks": _integer,
    "ad_orders": _integer,
    "ad_gmv": _decimal,
    "price": _decimal,
    "stock": _integer,
}


def import_report(session: Session, *, shop_id: int, path: Path) -> ImportSummary:
    shop = session.get(Shop, shop_id)
    if shop is None:
        raise LookupError("店铺不存在")

    headers, rows = _iter_rows(path)
    mapping = detect_columns(headers)
    if "sku_id" not in mapping:
        raise ValueError("无法识别 SKU ID 列，请确认报表包含 SKU ID/规格ID 等字段")
    if "date" not in mapping:
        raise ValueError("无法识别日期列，请确认报表包含 日期/统计日期 等字段")

    summary = ImportSummary()
    for row in rows:
        summary.rows_total += 1
        metric_date = _date(_lookup(row, mapping, "date"))
        sku_external_id = _text(_lookup(row, mapping, "sku_id"))
        if not metric_date or not sku_external_id:
            summary.rows_skipped += 1
            continue

        goods_id = _text(_lookup(row, mapping, "goods_id")) or f"report-{sku_external_id}"
        product = _get_or_create_product(
            session, shop_id=shop_id, goods_id=goods_id, summary=summary
        )
        sku = _get_or_create_sku(
            session, product=product, platform_sku_id=sku_external_id, summary=summary
        )

        metric = session.scalar(
            select(SkuDailyMetric).where(
                SkuDailyMetric.sku_id == sku.id,
                SkuDailyMetric.metric_date == metric_date,
            )
        )
        if metric is None:
            metric = SkuDailyMetric(
                metric_date=metric_date,
                shop_id=shop_id,
                product_id=product.id,
                sku_id=sku.id,
            )
            session.add(metric)
            summary.metrics_created += 1
        else:
            summary.metrics_updated += 1

        for field, converter in METRIC_FIELDS.items():
            if field not in mapping:
                continue
            parsed = converter(_lookup(row, mapping, field))
            if parsed is not None:
                setattr(metric, field, parsed)

        if metric.price is not None:
            sku.price = metric.price
        if metric.stock is not None:
            sku.stock = metric.stock

        summary.rows_imported += 1
        summary.min_date = metric_date if summary.min_date is None else min(summary.min_date, metric_date)
        summary.max_date = metric_date if summary.max_date is None else max(summary.max_date, metric_date)

    if summary.rows_imported == 0:
        raise ValueError("报表中没有可导入的有效行，请检查日期和 SKU ID")
    return summary
