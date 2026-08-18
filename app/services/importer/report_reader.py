from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


@dataclass(slots=True)
class ReportPreview:
    headers: list[str]
    rows: list[dict[str, object]]
    detected_fields: dict[str, str]


FIELD_ALIASES: dict[str, tuple[str, ...]] = {
    "date": ("日期", "时间", "统计日期", "date"),
    "goods_id": ("商品id", "商品ID", "goods_id", "商品编号"),
    "sku_id": ("sku_id", "SKU ID", "skuid", "规格ID", "sku编号"),
    "impression": ("曝光量", "曝光", "展现量", "impression"),
    "visitors": ("访客数", "访客", "uv", "visitors"),
    "clicks": ("点击量", "点击数", "点击", "clicks"),
    "order_count": ("订单数", "支付订单数", "成交订单数", "orders"),
    "sales_qty": ("销量", "支付件数", "成交件数", "sales_qty"),
    "gmv": ("成交金额", "支付金额", "gmv", "销售额"),
    "refund_count": ("退款订单数", "退款数", "refund_count"),
    "refund_amount": ("退款金额", "refund_amount"),
    "ad_cost": ("消耗", "推广消耗", "广告花费", "ad_cost"),
    "ad_clicks": ("推广点击", "广告点击", "ad_clicks"),
    "ad_orders": ("推广订单", "广告订单", "ad_orders"),
    "ad_gmv": ("推广成交金额", "广告成交金额", "ad_gmv"),
    "price": ("价格", "SKU价格", "商品价格", "price"),
    "stock": ("库存", "可售库存", "SKU库存", "stock"),
}


def _normalize(value: object) -> str:
    return str(value or "").strip().lower().replace(" ", "")


def detect_columns(headers: Iterable[object]) -> dict[str, str]:
    original = [str(h or "").strip() for h in headers]
    normalized = {_normalize(h): h for h in original if h}
    result: dict[str, str] = {}
    for field, aliases in FIELD_ALIASES.items():
        for alias in aliases:
            key = _normalize(alias)
            if key in normalized:
                result[field] = normalized[key]
                break
    return result


def preview_report(path: Path, limit: int = 20) -> ReportPreview:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            headers = list(reader.fieldnames or [])
            rows = [dict(row) for _, row in zip(range(limit), reader)]
    elif suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        iterator = sheet.iter_rows(values_only=True)
        headers = [str(value or "").strip() for value in next(iterator, ())]
        rows = []
        for _, values in zip(range(limit), iterator):
            rows.append({header: value for header, value in zip(headers, values)})
        workbook.close()
    else:
        raise ValueError("仅支持 CSV、XLSX、XLSM 报表")

    return ReportPreview(headers=headers, rows=rows, detected_fields=detect_columns(headers))
